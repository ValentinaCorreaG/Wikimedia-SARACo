"""
reports/generators/base.py
Abstract base class for all report generators
"""
from abc import ABC, abstractmethod
import pandas as pd
from io import BytesIO
from typing import BinaryIO, Dict, Any
from django.db import models
from datetime import datetime, date


class BaseReportGenerator(ABC):
    """
    Abstract base class for all report generators.
    Provides common functionality for generating Excel reports from Django models.
    """
    
    # Fields to exclude from automatic extraction
    EXCLUDED_FIELDS = ['id']
    
    def __init__(self, instance):
        """
        Initialize the generator with a model instance
        
        Args:
            instance: Django model instance to generate report from
        """
        self.instance = instance
        self.df = None
    
    def get_field_value(self, field) -> Any:
        """
        Extract value from a model field, handling different field types.
        
        Args:
            field: Django model field
            
        Returns:
            Formatted value appropriate for Excel
        """
        field_name = field.name
        value = getattr(self.instance, field_name)
        
        # Handle None/empty values
        if value is None or value == '':
            return ''
        
        # Handle ForeignKey relationships
        if isinstance(field, models.ForeignKey):
            return self.get_foreign_key_display(field_name, value)
        
        # Handle choice fields
        if field.choices:
            return self.get_choice_display(field_name)
        
        # Handle date/datetime fields
        if isinstance(field, (models.DateField, models.DateTimeField)):
            return self.format_date(value, isinstance(field, models.DateTimeField))
        
        # Handle boolean fields
        if isinstance(field, models.BooleanField):
            return 'Sí' if value else 'No'
        
        # Handle numeric fields
        if isinstance(field, (models.IntegerField, models.FloatField, models.DecimalField)):
            return value
        
        # Default: return as string
        return str(value)
    
    def get_foreign_key_display(self, field_name: str, related_instance) -> str:
        """
        Get display value for a ForeignKey field.
        Override this method to customize foreign key display logic.
        
        Args:
            field_name: Name of the foreign key field
            related_instance: Related model instance
            
        Returns:
            String representation of the related object
        """
        if related_instance is None:
            return ''
        
        # Try to get 'nombre' field first (common in Spanish models)
        if hasattr(related_instance, 'nombre'):
            return related_instance.nombre
        
        # Try 'name' field
        if hasattr(related_instance, 'name'):
            return related_instance.name
        
        # Fallback to __str__
        return str(related_instance)
    
    def get_choice_display(self, field_name: str) -> str:
        """
        Get display value for a field with choices.
        
        Args:
            field_name: Name of the field with choices
            
        Returns:
            Human-readable choice value
        """
        get_display_method = f'get_{field_name}_display'
        if hasattr(self.instance, get_display_method):
            return getattr(self.instance, get_display_method)()
        return getattr(self.instance, field_name, '')
    
    def format_date(self, date_value, include_time: bool = False) -> str:
        """
        Format date or datetime value for Excel display.
        
        Args:
            date_value: datetime or date object
            include_time: Whether to include time in the format
            
        Returns:
            Formatted date string
        """
        if date_value is None:
            return ''
        
        if include_time:
            return date_value.strftime('%Y-%m-%d %H:%M:%S')
        else:
            if isinstance(date_value, datetime):
                return date_value.date().strftime('%Y-%m-%d')
            return date_value.strftime('%Y-%m-%d')
    
    def get_field_label(self, field) -> str:
        """
        Get the verbose name (label) for a field.
        
        Args:
            field: Django model field
            
        Returns:
            Verbose name of the field
        """
        return field.verbose_name if hasattr(field, 'verbose_name') else field.name.replace('_', ' ').title()
    
    def extract_model_data(self) -> Dict[str, Any]:
        """
        Automatically extract all fields from the model instance.
        Creates a dictionary with verbose_name as key and formatted value.
        
        Returns:
            Dictionary of field labels and values
        """
        data = {}
        
        # Get all fields from the model
        fields = self.instance._meta.get_fields()
        
        for field in fields:
            # Skip excluded fields and reverse relations
            if field.name in self.EXCLUDED_FIELDS:
                continue
            
            # Skip reverse ForeignKey and ManyToMany relations
            if field.one_to_many or field.many_to_many:
                continue
            
            # Get label and value
            label = self.get_field_label(field)
            value = self.get_field_value(field)
            
            data[label] = value
        
        return data
    
    def _get_report_prefix(self) -> str:
        """
        Determine a human-friendly report prefix based on the model.
        Subclasses may override for a custom static prefix.
        """
        # Prefer Django verbose_name if available
        try:
            verbose_name = self.instance._meta.verbose_name
            if isinstance(verbose_name, str) and verbose_name.strip():
                return str(verbose_name).title().replace(' ', '_')
        except Exception:
            pass
        # Fallback to class name
        return self.instance.__class__.__name__
    
    def _get_primary_name_for_filename(self) -> str:
        """
        Get a representative name string from the instance to include in filenames.
        Tries common field names before falling back to __str__.
        """
        candidate_fields = ['nombre', 'name', 'titulo', 'title']
        for field_name in candidate_fields:
            if hasattr(self.instance, field_name):
                value = getattr(self.instance, field_name)
                if value:
                    return str(value)
        return str(self.instance)
    
    def _get_primary_date_for_filename(self):
        """
        Get a representative date/datetime from the instance for the filename.
        Tries common field names; returns today's date if not found.
        """
        candidate_fields = ['fecha', 'date', 'created', 'updated', 'created_at', 'updated_at']
        for field_name in candidate_fields:
            if hasattr(self.instance, field_name):
                value = getattr(self.instance, field_name)
                if isinstance(value, (datetime, date)):
                    return value
        return datetime.now()
    
    @abstractmethod
    def prepare_data(self):
        """
        Convert model instance to pandas DataFrame.
        Must set self.df with the prepared data.
        
        Should be implemented by subclasses to handle model-specific logic.
        """
        pass
    
    def get_filename(self) -> str:
        """
        Generate appropriate filename for the report, with a sensible default:
        {Prefix}_{SanitizedName}_{YYYYMMDD}.xlsx
        
        Returns:
            str: Filename with .xlsx extension
        """
        # Build prefix
        prefix = self._get_report_prefix()
        
        # Build safe name
        raw_name = self._get_primary_name_for_filename()
        safe_name = "".join(
            c for c in raw_name
            if c.isalnum() or c in (' ', '_', '-')
        ).strip()
        safe_name = safe_name.replace(' ', '_')[:50]
        
        # Build date segment
        date_value = self._get_primary_date_for_filename()
        date_str = self.format_date(date_value).replace('-', '')
        
        return f"{prefix}_{safe_name}_{date_str}.xlsx"
    
    def apply_formatting(self, writer):
        """
        Apply Excel formatting to the workbook.
        Provides a clean, professional default formatting.
        Can be overridden by subclasses for custom formatting.
        
        Args:
            writer: pandas ExcelWriter object
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        worksheet = writer.sheets['Datos']
        
        # Define default color scheme
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        data_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Format header row
        for col_num, column in enumerate(self.df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Format data rows
        for row_num in range(2, worksheet.max_row + 1):
            for col_num in range(1, len(self.df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.fill = data_fill
        
        # Adjust column widths based on content
        for col_num, column in enumerate(self.df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate width based on column name and content
            max_length = len(str(column))  # Start with header length
            
            # Check data length (sample first few rows for performance)
            for row_num in range(min(5, len(self.df))):
                cell_value = str(self.df.iloc[row_num, col_num - 1])
                max_length = max(max_length, len(cell_value))
            
            # Set width with reasonable bounds (min 15, max 50)
            adjusted_width = min(max(max_length + 2, 15), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights
        worksheet.row_dimensions[1].height = 35  # Header
        for row_num in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = 30  # Data rows
    
    def generate_excel(self) -> BinaryIO:
        """
        Generate Excel file and return as BytesIO object.
        This is the main method to call from views.
        
        Returns:
            BytesIO: Excel file in memory
            
        Raises:
            ValueError: If prepare_data() hasn't set self.df
        """
        # Prepare the data
        self.prepare_data()
        
        # Validate that DataFrame was created
        if self.df is None:
            raise ValueError("prepare_data() must set self.df")
        
        # Create BytesIO buffer
        buffer = BytesIO()
        
        # Write to Excel
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False, sheet_name='Datos')
            
            # Apply custom formatting if implemented
            self.apply_formatting(writer)
        
        # Reset buffer position to beginning
        buffer.seek(0)
        
        return buffer
    
    def generate_and_save(self, filepath: str):
        """
        Generate Excel file and save to disk.
        Useful for testing or batch generation.
        
        Args:
            filepath: Full path where to save the file
        """
        self.prepare_data()
        
        if self.df is None:
            raise ValueError("prepare_data() must set self.df")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False, sheet_name='Datos')
            self.apply_formatting(writer)
    
    def get_dataframe(self) -> pd.DataFrame:
        """
        Get the prepared DataFrame without generating Excel.
        Useful for testing or further processing.
        
        Returns:
            pd.DataFrame: Prepared data
        """
        self.prepare_data()
        return self.df
    
    def validate_instance(self):
        """
        Validate that the instance has required data.
        Can be overridden by subclasses for model-specific validation.
        
        Returns:
            tuple: (is_valid, error_message)
        """
        if not self.instance:
            return False, "No instance provided"
        
        return True, None